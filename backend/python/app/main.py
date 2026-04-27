from __future__ import annotations

import json
import os
import re
import hashlib
from copy import copy
from contextlib import contextmanager, ExitStack
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Literal

import fitz
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Color, Font, PatternFill
from PIL import Image
from pydantic import BaseModel, Field
try:
    import fcntl
except Exception:  # pragma: no cover
    fcntl = None

APP_NAME = "excel-edit-backend"
PYTHON_BACKEND_ROOT = Path(__file__).resolve().parents[1]
PROJECT_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_APP_DATA_ROOT = Path.home() / "app-data" / "company-search"
DEFAULT_STORAGE_ROOT = Path(
    os.getenv("EXCEL_EDIT_STORAGE_ROOT", str(PYTHON_BACKEND_ROOT / "data"))
).resolve()
DEFAULT_ARCHIVE_ROOT = (DEFAULT_STORAGE_ROOT / "archive").resolve()

RELATIVE_OFFSETS = {
    "회사명": -2,
    "대표자": -1,
    "사업자번호": 0,
    "지역": 1,
    "시공능력": 2,
    "3년간 실적액": 3,
    "5년간 실적액": 4,
    "부채비율": 5,
    "유동비율": 6,
    "영업기간공사업등록일": 7,
    "신용평가": 8,
    "여성기업": 9,
    "중소기업": 10,
    "일자리창출실적": 11,
    "시공품질평가": 12,
    "비고": 13,
}

COLUMN_MAP = {
    "상호": "회사명",
    "대표자": "대표자",
    "사업자등록번호": "사업자번호",
    "지역": "지역",
    "시평액": "시공능력",
    "3년실적": "3년간 실적액",
    "5년실적": "5년간 실적액",
    "부채비율": "부채비율",
    "유동비율": "유동비율",
    "영업기간": "영업기간공사업등록일",
    "신용평가": "신용평가",
    "여성기업": "여성기업",
    "중소기업": "중소기업",
    "일자리창출실적": "일자리창출실적",
    "시공품질평가": "시공품질평가",
    "비고": "비고",
}

FORM_KEY_TO_KR = {
    "companyName": "상호",
    "managerName": "대표자",
    "bizNo": "사업자등록번호",
    "region": "지역",
    "sipyung": "시평액",
    "perf3y": "3년실적",
    "perf5y": "5년실적",
    "debtRatio": "부채비율",
    "currentRatio": "유동비율",
    "bizYears": "영업기간",
    "creditText": "신용평가",
    "womenOwned": "여성기업",
    "smallBusiness": "중소기업",
    "jobCreation": "일자리창출실적",
    "qualityEval": "시공품질평가",
    "note": "비고",
}

KR_TO_FORM_KEY = {value: key for key, value in FORM_KEY_TO_KR.items()}

FILE_TYPE_TO_DB_KEY = {
    "전기경영상태": "전기",
    "통신경영상태": "통신",
    "소방경영상태": "소방",
}

DB_ENV = {
    "전기": "EXCEL_EDIT_DB_PATH_ELECTRIC",
    "통신": "EXCEL_EDIT_DB_PATH_COMMUNICATION",
    "소방": "EXCEL_EDIT_DB_PATH_FIRE",
}

DB_DATASET_FILE = {
    "전기": "eung.xlsx",
    "통신": "tongsin.xlsx",
    "소방": "sobang.xlsx",
}

REGION_FOLDER_ALIASES = {
    "서울특별시": "서울",
    "서울": "서울",
    "부산광역시": "부산",
    "부산": "부산",
    "대구광역시": "대구",
    "대구": "대구",
    "인천광역시": "인천",
    "인천": "인천",
    "광주광역시": "광주",
    "광주": "광주",
    "대전광역시": "대전",
    "대전": "대전",
    "울산광역시": "울산",
    "울산": "울산",
    "세종특별자치시": "세종",
    "세종": "세종",
    "경기도": "경기",
    "경기": "경기",
    "강원특별자치도": "강원",
    "강원도": "강원",
    "강원": "강원",
    "충청북도": "충북",
    "충북": "충북",
    "충청남도": "충남",
    "충남": "충남",
    "전북특별자치도": "전북",
    "전라북도": "전북",
    "전북": "전북",
    "전라남도": "전남",
    "전남": "전남",
    "경상북도": "경북",
    "경북": "경북",
    "경상남도": "경남",
    "경남": "경남",
    "제주특별자치도": "제주",
    "제주도": "제주",
    "제주": "제주",
}

RATIO_THRESHOLDS = {
    "전기": {
        "부채비율": {"max": 62.02},
        "유동비율": {"min": 213.87},
    },
    "통신": {
        "부채비율": {"max": 62.01},
        "유동비율": {"min": 210.09},
    },
    "소방": {
        "부채비율": {"max": 55.04},
        "유동비율": {"min": 208.98},
    },
}

GREEN_THEME = Color(type="theme", theme=6, tint=0.7999816888943144)
BLUE_THEME = Color(type="theme", theme=3, tint=0.7999816888943144)
GREEN_FILL = PatternFill(fgColor=GREEN_THEME, fill_type="solid")
BLUE_FILL = PatternFill(fgColor=BLUE_THEME, fill_type="solid")
NO_FILL = PatternFill(fill_type=None)
GREY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
DEFAULT_FONT = Font(color="000000", bold=False, size=9)
HIGHLIGHT_FONT = Font(color="FF0000", bold=True, size=9)
BOLD_FONT = Font(bold=True, size=12)


class JobRequest(BaseModel):
    fileType: Literal["전기경영상태", "통신경영상태", "소방경영상태", "신용평가"] | str = Field(default="전기경영상태")
    excelPath: str = Field(default="")
    dryRun: bool = Field(default=False)


class ApiResponse(BaseModel):
    success: bool = True
    message: str = ""
    data: dict = Field(default_factory=dict)


class LookupRequest(BaseModel):
    fileType: str = Field(default="전기경영상태")
    bizNo: str = Field(default="")


class SaveRequest(BaseModel):
    fileType: str = Field(default="전기경영상태")
    bizNo: str = Field(default="")
    data: dict = Field(default_factory=dict)
    saveMode: Literal["normal", "archive_only"] | str = Field(default="normal")
    expectedVersion: str = Field(default="")


class DeleteRequest(BaseModel):
    fileType: str = Field(default="전기경영상태")
    bizNo: str = Field(default="")
    expectedVersion: str = Field(default="")


app = FastAPI(title=APP_NAME, version="0.1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://127.0.0.1:5173",
        "http://localhost:4173",
        "http://127.0.0.1:4173",
    ],
    allow_origin_regex=r"^http://(localhost|127\.0\.0\.1|10\.\d+\.\d+\.\d+|172\.(1[6-9]|2\d|3[0-1])\.\d+\.\d+|192\.168\.\d+\.\d+)(:\d+)?$",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["x-pdf-page-count", "x-pdf-page-number"],
)


def _now_stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")


def _safe_filename(name: str) -> str:
    keep = "._-()[]{}"
    cleaned = "".join(ch for ch in name if ch.isalnum() or ch in keep)
    return cleaned or f"uploaded_{_now_stamp()}"


def _normalize_biz_no(raw: str) -> str:
    return re.sub(r"\D", "", str(raw or ""))


def _format_biz_no(raw: str) -> str:
    digits = _normalize_biz_no(raw)[:10]
    if len(digits) <= 3:
        return digits
    if len(digits) <= 5:
        return f"{digits[:3]}-{digits[3:]}"
    return f"{digits[:3]}-{digits[3:5]}-{digits[5:]}"


def _clean_numeric_text(raw: str) -> str:
    return re.sub(r"[^0-9.]", "", str(raw or ""))


def _to_int_or_none(raw: str) -> int | None:
    digits = re.sub(r"\D", "", str(raw or ""))
    if not digits:
        return None
    return int(digits)


def _to_percent_or_none(raw: str) -> float | None:
    text = _clean_numeric_text(raw)
    if not text:
        return None
    return float(text)


def _format_amount_for_display(value: object) -> str:
    if value is None or str(value).strip() == "":
        return ""
    try:
        numeric = int(float(str(value).replace(",", "")))
        return f"{numeric:,}"
    except (TypeError, ValueError):
        return str(value)


def _format_amount_for_form_thousand_unit(value: object) -> str:
    if value is None or str(value).strip() == "":
        return ""
    try:
        numeric = int(float(str(value).replace(",", "")))
        return f"{int(round(numeric / 1000)):,}"
    except (TypeError, ValueError):
        return ""


def _format_ratio_for_display(value: object) -> str:
    if value is None or str(value).strip() == "":
        return ""
    try:
        if isinstance(value, (int, float)):
            numeric = float(value)
            if numeric <= 1:
                numeric *= 100
            return f"{numeric:.2f}%"
        text = str(value).strip()
        if text.endswith("%"):
            return f"{float(text[:-1]):.2f}%"
        numeric = float(text)
        if numeric <= 1:
            numeric *= 100
        return f"{numeric:.2f}%"
    except (TypeError, ValueError):
        return str(value)


def _resolve_db_paths() -> dict[str, str]:
    paths: dict[str, str] = {}
    dataset_root = _resolve_dataset_root()
    for key, file_name in DB_DATASET_FILE.items():
        candidate = (dataset_root / file_name).resolve()
        if candidate.exists():
            paths[key] = str(candidate)
    return paths


def _resolve_dataset_root() -> Path:
    explicit = (
        os.getenv("DATASET_UPLOAD_DIR", "").strip()
        or os.getenv("COMPANY_SEARCH_DATASET_DIR", "").strip()
    )
    if explicit:
        return Path(explicit).resolve()

    app_data = os.getenv("COMPANY_SEARCH_APP_DATA_ROOT", "").strip()
    app_root = Path(app_data).resolve() if app_data else DEFAULT_APP_DATA_ROOT.resolve()
    return (app_root / "uploads" / "master-files").resolve()


def _load_legacy_config() -> dict:
    candidates = [
        PROJECT_ROOT / "ocr_config.json",
        PROJECT_ROOT / "clone" / "excel_modifi" / "ocr_config.json",
    ]
    for candidate in candidates:
        if not candidate.exists():
            continue
        try:
            return json.loads(candidate.read_text(encoding="utf-8"))
        except Exception:
            continue
    return {}


def _resolve_archive_root() -> Path:
    env_value = os.getenv("EXCEL_EDIT_ARCHIVE_ROOT", "").strip()
    if env_value:
        return Path(env_value).resolve()

    legacy = _load_legacy_config()
    legacy_archive = str(legacy.get("archive_path", "")).strip()
    if legacy_archive:
        return Path(legacy_archive).resolve()

    return DEFAULT_ARCHIVE_ROOT


@contextmanager
def _exclusive_excel_lock(excel_path: str):
    lock_path = Path(f"{excel_path}.lock")
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    with open(lock_path, "a+b") as lock_file:
        if fcntl:
            fcntl.flock(lock_file.fileno(), fcntl.LOCK_EX)
        try:
            yield
        finally:
            if fcntl:
                fcntl.flock(lock_file.fileno(), fcntl.LOCK_UN)


@contextmanager
def _exclusive_excel_locks(excel_paths: list[str]):
    unique_paths = sorted({str(path) for path in excel_paths if str(path).strip()})
    with ExitStack() as stack:
        for path in unique_paths:
            stack.enter_context(_exclusive_excel_lock(path))
        yield


def _build_company_fingerprint(db_type: str, excel_path: str, sheet_name: str, row: int, col: int, raw: dict) -> str:
    payload = {
        "dbType": db_type,
        "excelPath": str(excel_path),
        "sheetName": str(sheet_name),
        "row": int(row),
        "col": int(col),
        "values": {key: str(raw.get(key) or "") for key in sorted(COLUMN_MAP.keys())},
    }
    digest = hashlib.sha256(json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8")).hexdigest()
    return digest[:24]


def _build_found_version(db_type: str, excel_path: str, biz_no: str, sheet_name: str, row: int, col: int, raw: dict) -> str:
    biz = _normalize_biz_no(biz_no)
    fingerprint = _build_company_fingerprint(db_type, excel_path, sheet_name, row, col, raw)
    return f"found|{db_type}|{biz}|{fingerprint}"


def _build_not_found_version(file_type: str, biz_no: str) -> str:
    return f"nf|{file_type}|{_normalize_biz_no(biz_no)}"


def _validate_expected_version_for_save(request: SaveRequest, db_paths: dict[str, str], expected_version: str) -> None:
    token = str(expected_version or "").strip()
    if not token:
        return

    parts = token.split("|")
    if len(parts) < 3:
        raise HTTPException(status_code=409, detail="저장 버전 정보가 올바르지 않습니다. 다시 조회 후 저장하세요.")

    normalized_biz = _normalize_biz_no(request.bizNo)
    token_kind = parts[0]
    token_file_type_or_db = parts[1]
    token_biz = parts[2]

    if token_biz and normalized_biz and token_biz != normalized_biz:
        raise HTTPException(status_code=409, detail="사업자번호 기준 데이터가 변경되었습니다. 다시 조회 후 저장하세요.")

    if token_kind == "nf":
        if token_file_type_or_db != request.fileType:
            raise HTTPException(status_code=409, detail="조회 기준 자료종류가 변경되었습니다. 다시 조회 후 저장하세요.")

        if request.fileType == "신용평가":
            for _, path in db_paths.items():
                if not path or not Path(path).exists():
                    continue
                if _find_company_position(path, request.bizNo):
                    raise HTTPException(status_code=409, detail="다른 사용자가 업체를 추가했습니다. 다시 조회 후 저장하세요.")
            return

        db_key = FILE_TYPE_TO_DB_KEY.get(request.fileType)
        target_path = db_paths.get(db_key, "") if db_key else ""
        if target_path and Path(target_path).exists() and _find_company_position(target_path, request.bizNo):
            raise HTTPException(status_code=409, detail="다른 사용자가 업체를 추가했습니다. 다시 조회 후 저장하세요.")
        return

    if token_kind != "found" or len(parts) < 4:
        raise HTTPException(status_code=409, detail="저장 버전 정보가 올바르지 않습니다. 다시 조회 후 저장하세요.")

    db_type = token_file_type_or_db
    expected_fp = parts[3]
    if request.fileType != "신용평가":
        target_db = FILE_TYPE_TO_DB_KEY.get(request.fileType)
        if target_db and db_type != target_db:
            raise HTTPException(status_code=409, detail="조회한 자료종류가 변경되었습니다. 다시 조회 후 저장하세요.")

    excel_path = db_paths.get(db_type, "")
    if not excel_path or not Path(excel_path).exists():
        raise HTTPException(status_code=409, detail="조회한 원본 DB가 변경되었습니다. 다시 조회 후 저장하세요.")

    position = _find_company_position(excel_path, request.bizNo)
    if not position:
        raise HTTPException(status_code=409, detail="조회한 업체 데이터가 변경되었습니다. 다시 조회 후 저장하세요.")

    sheet_name, row, col = position
    raw = _extract_company_data(excel_path, sheet_name, row, col)
    current_fp = _build_company_fingerprint(db_type, excel_path, sheet_name, row, col, raw)
    if current_fp != expected_fp:
        raise HTTPException(status_code=409, detail="다른 사용자가 먼저 수정했습니다. 다시 조회 후 저장하세요.")


def _find_company_position(excel_path: str, biz_no: str) -> tuple[str, int, int] | None:
    normalized_target = _normalize_biz_no(biz_no)
    if not normalized_target:
        return None
    workbook = load_workbook(filename=excel_path, data_only=False)
    try:
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    normalized_cell = _normalize_biz_no(str(cell.value))
                    if normalized_cell and normalized_cell == normalized_target:
                        return sheet_name, cell.row, cell.column
        return None
    finally:
        workbook.close()


def _extract_company_data(excel_path: str, sheet_name: str, row: int, col: int) -> dict:
    workbook = load_workbook(filename=excel_path, data_only=False)
    try:
        sheet = workbook[sheet_name]
        result: dict[str, object] = {}
        for kr_key, excel_label in COLUMN_MAP.items():
            offset = RELATIVE_OFFSETS.get(excel_label)
            if offset is None:
                continue
            target_row = row + offset
            if 1 <= target_row <= sheet.max_row and 1 <= col <= sheet.max_column:
                cell = _resolve_merged_cell(sheet, target_row, col)
                result[kr_key] = cell.value
            else:
                result[kr_key] = ""
        result["지역"] = sheet_name
        return result
    finally:
        workbook.close()


def _resolve_fill_color_hex(cell) -> str:
    color_hex = "#FFFFFF"
    fill = getattr(cell, "fill", None)
    fg_color = getattr(fill, "fgColor", None) if fill else None
    if not fg_color:
        return color_hex

    color_type = getattr(fg_color, "type", None)
    if color_type == "theme":
        theme = getattr(fg_color, "theme", None)
        if theme == 6:
            return "#E2EFDA"
        if theme == 3:
            return "#DDEBF7"
        return color_hex

    if color_type == "rgb":
        rgb = str(getattr(fg_color, "rgb", "") or "").strip()
        if len(rgb) == 8 and rgb.upper().startswith("FF"):
            return f"#{rgb[2:].upper()}"
        if len(rgb) == 6:
            return f"#{rgb.upper()}"

    return color_hex


def _extract_company_cells(excel_path: str, sheet_name: str, row: int, col: int) -> dict:
    workbook = load_workbook(filename=excel_path, data_only=False)
    try:
        sheet = workbook[sheet_name]
        result: dict[str, dict] = {}
        for kr_key, excel_label in COLUMN_MAP.items():
            offset = RELATIVE_OFFSETS.get(excel_label)
            if offset is None:
                continue
            target_row = row + offset
            value = ""
            color = "#FFFFFF"
            if 1 <= target_row <= sheet.max_row and 1 <= col <= sheet.max_column:
                cell = _resolve_merged_cell(sheet, target_row, col)
                value = cell.value
                color = _resolve_fill_color_hex(cell)
            result[kr_key] = {"value": value, "color": color}
        result["지역"] = {"value": sheet_name, "color": "#FFFFFF"}
        return result
    finally:
        workbook.close()


def _resolve_merged_cell(sheet, row: int, col: int):
    cell = sheet.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        return cell

    for merged_range in sheet.merged_cells.ranges:
        if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
            return sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
    return cell


def _normalize_label(value: object) -> str:
    return re.sub(r"\s+", "", str(value or ""))


def _is_empty_company_slot(value: object) -> bool:
    text = str(value or "").strip()
    return text in {"", "-", "—", "N/A", "없음"}


def _normalize_region_folder_name(region: str) -> str:
    raw = str(region or "").strip()
    if not raw:
        return "기타"

    token = raw.split()[0]
    if token in REGION_FOLDER_ALIASES:
        return REGION_FOLDER_ALIASES[token]

    for long_name, short_name in sorted(REGION_FOLDER_ALIASES.items(), key=lambda item: len(item[0]), reverse=True):
        if token.startswith(long_name):
            return short_name
    return token


def _parse_expiry_date(raw: str):
    for fmt in ("%y.%m.%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _parse_pdf_page_selection(selection: str, max_page: int) -> list[int]:
    text = str(selection or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="내보낼 페이지 범위를 입력하세요.")

    pages: set[int] = set()
    for token in [part.strip() for part in text.split(",") if part.strip()]:
        if "-" in token:
            bounds = [part.strip() for part in token.split("-", 1)]
            if len(bounds) != 2 or not bounds[0].isdigit() or not bounds[1].isdigit():
                raise HTTPException(status_code=400, detail=f"잘못된 페이지 범위: {token}")
            start = int(bounds[0])
            end = int(bounds[1])
            if start > end:
                start, end = end, start
            if start < 1 or end > max_page:
                raise HTTPException(status_code=400, detail=f"페이지 범위를 벗어났습니다: {token}")
            pages.update(range(start - 1, end))
            continue

        if not token.isdigit():
            raise HTTPException(status_code=400, detail=f"잘못된 페이지 번호: {token}")
        page_no = int(token)
        if page_no < 1 or page_no > max_page:
            raise HTTPException(status_code=400, detail=f"페이지 범위를 벗어났습니다: {token}")
        pages.add(page_no - 1)

    if not pages:
        raise HTTPException(status_code=400, detail="내보낼 페이지를 찾지 못했습니다.")
    return sorted(pages)


def _is_theme_color(fill, theme: int) -> bool:
    fg = getattr(fill, "fgColor", None)
    return (
        getattr(fill, "fill_type", None) == "solid"
        and getattr(fg, "type", None) == "theme"
        and getattr(fg, "theme", None) == theme
    )


def _apply_management_cell_style(cell, kr_key: str) -> None:
    if kr_key not in {"상호", "신용평가"}:
        cell.fill = GREEN_FILL


def _update_cell_value_by_key(cell, kr_key: str, raw_value: str) -> tuple[bool, float | None]:
    value_text = str(raw_value or "").strip()
    if value_text == "":
        return False, None

    if kr_key in {"시평액", "3년실적", "5년실적"}:
        numeric = _to_int_or_none(value_text)
        if numeric is None:
            return False, None
        cell.value = numeric * 1000
        return True, None

    if kr_key in {"부채비율", "유동비율"}:
        numeric = _to_percent_or_none(value_text)
        if numeric is None:
            return False, None
        cell.value = numeric / 100.0
        cell.number_format = "0.00%"
        return True, numeric

    cell.value = value_text
    return True, None


def _resolve_sheet_name_for_new_company(workbook, requested_region: str) -> str | None:
    wanted = str(requested_region or "").strip()
    if not wanted:
        return None
    if wanted in workbook.sheetnames:
        return wanted

    # GUI 입력 습관 대응: "서울 ..." 같이 접두어만 입력되는 경우를 허용한다.
    wanted_head = wanted.split()[0]
    for sheet_name in workbook.sheetnames:
        if sheet_name.startswith(wanted_head):
            return sheet_name
    return None


def _find_company_section_bounds(sheet, row: int) -> tuple[int, int] | None:
    company_label, remarks_label = "회사명", "비고"
    section_start = -1
    section_end = -1

    for scan_row in range(row, 0, -1):
        label_value = _normalize_label(sheet.cell(row=scan_row, column=1).value)
        if company_label in label_value:
            section_start = scan_row
            break

    if section_start == -1:
        return None

    for scan_row in range(row, sheet.max_row + 1):
        label_value = _normalize_label(sheet.cell(row=scan_row, column=1).value)
        if remarks_label in label_value:
            section_end = scan_row
            break

    if section_end == -1 or section_end < section_start:
        return None
    return section_start, section_end


def _add_new_company_data(excel_path: str, form_data: dict, db_type: str) -> dict:
    company_name = str(form_data.get("companyName") or "").strip()
    requested_region = str(form_data.get("region") or "").strip()
    if not company_name or not requested_region:
        raise HTTPException(status_code=400, detail="신규 업체 추가에는 상호와 지역(시트명)이 필요합니다.")

    workbook = load_workbook(filename=excel_path)
    try:
        sheet_name = _resolve_sheet_name_for_new_company(workbook, requested_region)
        if not sheet_name:
            raise HTTPException(status_code=400, detail=f"'{requested_region}'에 해당하는 시트를 찾지 못했습니다.")
        sheet = workbook[sheet_name]

        start_col, end_col = 2, 13
        company_label, remarks_label = "회사명", "비고"
        sections: list[tuple[int, int]] = []
        start_row = -1
        for row in range(1, sheet.max_row + 2):
            label_value = _normalize_label(sheet.cell(row=row, column=1).value)
            if company_label in label_value:
                start_row = row
            elif remarks_label in label_value and start_row != -1:
                sections.append((start_row, row))
                start_row = -1

        if not sections:
            raise HTTPException(status_code=400, detail=f"'{sheet_name}' 시트의 데이터 구간을 찾지 못했습니다.")

        target_row, target_col = -1, -1
        found_empty_slot = False
        for sec_start, sec_end in sections:
            company_row = next(
                (
                    r
                    for r in range(sec_start, sec_end + 1)
                    if company_label in _normalize_label(sheet.cell(row=r, column=1).value)
                ),
                -1,
            )
            if company_row == -1:
                continue
            for col in range(start_col, end_col + 1):
                cell = _resolve_merged_cell(sheet, company_row, col)
                if _is_empty_company_slot(cell.value):
                    target_row, target_col = sec_start, col
                    found_empty_slot = True
                    break
            if found_empty_slot:
                break

        if not found_empty_slot:
            last_start, last_end = sections[-1]
            section_height = last_end - last_start
            target_row = last_end + 1
            target_col = start_col

            for i in range(section_height + 1):
                source_row = last_start + i
                new_row = target_row + i
                if sheet.row_dimensions[source_row].height:
                    sheet.row_dimensions[new_row].height = sheet.row_dimensions[source_row].height

                for col in range(1, end_col + 1):
                    source_cell = sheet.cell(row=source_row, column=col)
                    new_cell = sheet.cell(row=new_row, column=col)
                    if source_cell.has_style:
                        new_cell.font = copy(source_cell.font)
                        new_cell.border = copy(source_cell.border)
                        new_cell.number_format = source_cell.number_format
                        new_cell.protection = copy(source_cell.protection)
                        new_cell.alignment = copy(source_cell.alignment)
                        new_cell.fill = copy(source_cell.fill) if col == 1 else copy(NO_FILL)
                    new_cell.value = source_cell.value if col == 1 else None

            sections.append((target_row, target_row + section_height))

        section_start = next((s for s, e in sections if s <= target_row <= e), -1)
        section_end = next((e for s, e in sections if s == section_start), sheet.max_row)
        labels_in_section: dict[str, int] = {}
        for row in range(section_start, section_end + 1):
            label_value = sheet.cell(row=row, column=1).value
            if label_value is None:
                continue
            labels_in_section[_normalize_label(label_value)] = row

        final_data: dict[str, str] = {"상호": company_name}
        for form_key, raw in form_data.items():
            kr_key = FORM_KEY_TO_KR.get(form_key)
            if not kr_key:
                continue
            final_data[kr_key] = str(raw or "")

        for kr_key, raw_value in final_data.items():
            excel_label = COLUMN_MAP.get(kr_key)
            if not excel_label:
                continue
            normalized_label = _normalize_label(excel_label)
            if normalized_label == "상호":
                final_label = "회사명"
            elif normalized_label == "시공능력":
                final_label = f"{db_type}시공능력"
            else:
                final_label = normalized_label

            row_to_insert = labels_in_section.get(final_label)
            if not row_to_insert:
                continue

            cell = _resolve_merged_cell(sheet, row_to_insert, target_col)
            cell.font = copy(DEFAULT_FONT)
            if final_label == "회사명":
                cell.fill = copy(GREY_FILL)
                cell.font = copy(BOLD_FONT)
            else:
                cell.fill = copy(GREEN_FILL)

            value_text = str(raw_value or "").strip()
            if value_text == "":
                continue
            _update_cell_value_by_key(cell, kr_key, value_text)

        workbook.save(excel_path)
        return {"sheetName": sheet_name, "companyName": company_name}
    finally:
        workbook.close()


def _build_lookup_payload(raw: dict, db_type: str, excel_path: str, raw_cells: dict | None = None) -> dict:
    company = {
        "companyName": str(raw.get("상호") or ""),
        "managerName": str(raw.get("대표자") or ""),
        "bizNo": _format_biz_no(str(raw.get("사업자등록번호") or "")),
        "region": str(raw.get("지역") or ""),
        "sipyung": _format_amount_for_display(raw.get("시평액")),
        "perf3y": _format_amount_for_display(raw.get("3년실적")),
        "perf5y": _format_amount_for_display(raw.get("5년실적")),
        "debtRatio": _format_ratio_for_display(raw.get("부채비율")),
        "currentRatio": _format_ratio_for_display(raw.get("유동비율")),
        "bizYears": str(raw.get("영업기간") or ""),
        "creditText": str(raw.get("신용평가") or ""),
        "womenOwned": str(raw.get("여성기업") or ""),
        "smallBusiness": str(raw.get("중소기업") or ""),
        "jobCreation": str(raw.get("일자리창출실적") or ""),
        "qualityEval": str(raw.get("시공품질평가") or ""),
        "note": str(raw.get("비고") or ""),
    }

    form_defaults = {
        "companyName": company["companyName"],
        "managerName": company["managerName"],
        "bizNo": company["bizNo"],
        "region": company["region"],
        "sipyung": _format_amount_for_form_thousand_unit(raw.get("시평액")),
        "perf3y": _format_amount_for_form_thousand_unit(raw.get("3년실적")),
        "perf5y": _format_amount_for_form_thousand_unit(raw.get("5년실적")),
        "debtRatio": company["debtRatio"],
        "currentRatio": company["currentRatio"],
        "bizYears": company["bizYears"],
        "womenOwned": company["womenOwned"],
        "smallBusiness": company["smallBusiness"],
        "jobCreation": company["jobCreation"],
        "qualityEval": company["qualityEval"],
        "note": company["note"],
    }

    color_map = {}
    if raw_cells:
        for kr_key, cell_info in raw_cells.items():
            form_key = KR_TO_FORM_KEY.get(kr_key)
            if not form_key:
                continue
            color_map[form_key] = str((cell_info or {}).get("color") or "#FFFFFF")

    return {
        "company": company,
        "formDefaults": form_defaults,
        "colorMap": color_map,
        "source": {
            "dbType": db_type,
            "excelPath": excel_path,
            "sheetName": company["region"],
        },
    }


def _update_management_data(excel_path: str, biz_no: str, form_data: dict, db_type: str) -> list[str]:
    position = _find_company_position(excel_path, biz_no)
    if not position:
        raise HTTPException(status_code=404, detail="해당 사업자번호의 업체를 찾을 수 없습니다.")

    sheet_name, target_row, target_col = position
    workbook = load_workbook(filename=excel_path)
    updated_labels: list[str] = []
    try:
        sheet = workbook[sheet_name]
        for form_key, raw_value in form_data.items():
            if form_key not in FORM_KEY_TO_KR:
                continue
            if form_key in {"creditGrade", "creditStartDate", "creditEndDate"}:
                continue
            kr_key = FORM_KEY_TO_KR[form_key]
            excel_label = COLUMN_MAP.get(kr_key)
            offset = RELATIVE_OFFSETS.get(excel_label or "")
            if excel_label is None or offset is None:
                continue

            update_row = target_row + offset
            if not (1 <= update_row <= sheet.max_row and 1 <= target_col <= sheet.max_column):
                continue

            cell = _resolve_merged_cell(sheet, update_row, target_col)
            _apply_management_cell_style(cell, kr_key)
            if kr_key != "상호":
                cell.font = copy(DEFAULT_FONT)

            value_text = str(raw_value or "").strip()
            updated, ratio_value = _update_cell_value_by_key(cell, kr_key, value_text)
            if not updated:
                continue
            updated_labels.append(excel_label)

            if kr_key in {"부채비율", "유동비율"} and ratio_value is not None:
                thresholds = RATIO_THRESHOLDS.get(db_type, {}).get(kr_key, {})
                if ("max" in thresholds and ratio_value > thresholds["max"]) or (
                    "min" in thresholds and ratio_value < thresholds["min"]
                ):
                    cell.font = copy(HIGHLIGHT_FONT)

        workbook.save(excel_path)
        return updated_labels
    finally:
        workbook.close()


def _delete_management_company(excel_path: str, biz_no: str, db_type: str) -> dict:
    position = _find_company_position(excel_path, biz_no)
    if not position:
        raise HTTPException(status_code=404, detail="삭제 대상 업체를 찾지 못했습니다.")

    sheet_name, target_row, target_col = position
    workbook = load_workbook(filename=excel_path)
    try:
        sheet = workbook[sheet_name]
        section_bounds = _find_company_section_bounds(sheet, target_row)
        if not section_bounds:
            raise HTTPException(status_code=400, detail="업체 블록 범위를 찾지 못했습니다.")

        section_start, section_end = section_bounds
        cleared_labels: list[str] = []

        for row in range(section_start, section_end + 1):
            label_value = sheet.cell(row=row, column=1).value
            normalized_label = _normalize_label(label_value)
            cell = _resolve_merged_cell(sheet, row, target_col)
            cell.value = None
            cell.fill = copy(NO_FILL)
            cell.font = copy(DEFAULT_FONT)
            cell.number_format = "General"
            if normalized_label:
                cleared_labels.append(str(label_value))

        workbook.save(excel_path)
        return {
            "dbType": db_type,
            "excelPath": excel_path,
            "sheetName": sheet_name,
            "sectionStart": section_start,
            "sectionEnd": section_end,
            "clearedLabels": cleared_labels,
        }
    finally:
        workbook.close()


def _build_credit_text(form_data: dict) -> str:
    grade = str(form_data.get("creditGrade") or "").strip()
    start = str(form_data.get("creditStartDate") or "").strip()
    end = str(form_data.get("creditEndDate") or "").strip()
    direct = str(form_data.get("creditText") or "").strip()
    if direct:
        return direct
    if not grade and not start and not end:
        return ""
    if not grade:
        return f"{start or '?'}~{end or '?'}"
    if not start and not end:
        return grade
    return f"{grade}\n({start or '?'}~{end or '?'})"


def _update_credit_data(db_paths: dict[str, str], biz_no: str, credit_text: str) -> list[dict]:
    results: list[dict] = []
    for db_type, excel_path in db_paths.items():
        if not Path(excel_path).exists():
            continue

        position = _find_company_position(excel_path, biz_no)
        if not position:
            results.append({"dbType": db_type, "updated": False, "reason": "not_found"})
            continue

        sheet_name, target_row, target_col = position
        workbook = load_workbook(filename=excel_path)
        try:
            sheet = workbook[sheet_name]
            offset = RELATIVE_OFFSETS["신용평가"]
            update_row = target_row + offset
            if 1 <= update_row <= sheet.max_row and 1 <= target_col <= sheet.max_column:
                cell = _resolve_merged_cell(sheet, update_row, target_col)
                cell.value = credit_text
                cell.fill = copy(GREEN_FILL)
                workbook.save(excel_path)
                results.append({"dbType": db_type, "updated": True, "sheetName": sheet_name})
            else:
                results.append({"dbType": db_type, "updated": False, "reason": "invalid_cell"})
        finally:
            workbook.close()
    return results


def _archive_uploaded_files(files: list[UploadFile], company_name: str, file_type: str, region: str) -> list[dict]:
    archived: list[dict] = []
    normalized_region = _normalize_region_folder_name(region)
    region_safe = re.sub(r'[<>:"/\\|?*]+', "", normalized_region).strip() or "기타"
    company_safe = re.sub(r'[<>:"/\\|?*]+', "", str(company_name or "업체")).strip() or "업체"
    company_safe = company_safe.replace("㈜", "(주)")

    target_dir = _resolve_archive_root() / region_safe
    target_dir.mkdir(parents=True, exist_ok=True)

    for idx, item in enumerate(files):
        ext = Path(item.filename or "").suffix or ""
        base_name = f"{company_safe}_{file_type}"
        file_name = f"{base_name}{ext}" if idx == 0 else f"{base_name}_{idx + 1}{ext}"
        target_path = target_dir / file_name

        if target_path.exists():
            target_path.unlink()

        content = item.file.read()
        target_path.write_bytes(content)
        archived.append({
            "originalName": item.filename or "",
            "savedName": file_name,
            "path": str(target_path),
            "size": len(content),
        })

    return archived


def _resolve_target_excel_paths_for_job(payload: JobRequest) -> list[tuple[str, str]]:
    db_paths = _resolve_db_paths()
    if payload.excelPath and Path(payload.excelPath).exists():
        return [("direct", payload.excelPath)]

    if payload.fileType == "신용평가":
        return [(db_type, path) for db_type, path in db_paths.items() if Path(path).exists()]

    db_key = FILE_TYPE_TO_DB_KEY.get(payload.fileType)
    if not db_key:
        raise HTTPException(status_code=400, detail=f"지원하지 않는 자료 종류입니다: {payload.fileType}")
    excel_path = db_paths.get(db_key, "")
    if not excel_path or not Path(excel_path).exists():
        raise HTTPException(status_code=400, detail=f"{db_key} DB 경로가 유효하지 않습니다.")
    return [(db_key, excel_path)]


def _batch_update_year_end_colors(excel_path: str, dry_run: bool = False) -> dict:
    workbook = load_workbook(filename=excel_path)
    update_count = 0
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(min_row=2):
                label_cell = row[0]
                if label_cell.value and "신용평가" in str(label_cell.value):
                    continue
                for cell in row[1:]:
                    value_text = str(cell.value or "").strip()
                    if value_text == "":
                        if getattr(cell.fill, "fill_type", None) is not None:
                            cell.fill = copy(NO_FILL)
                            update_count += 1
                        continue

                    if _is_theme_color(cell.fill, 6):
                        cell.fill = copy(BLUE_FILL)
                        update_count += 1
                    elif _is_theme_color(cell.fill, 3):
                        cell.fill = copy(NO_FILL)
                        update_count += 1

        if not dry_run:
            workbook.save(excel_path)
        return {"updatedCount": update_count, "dryRun": dry_run}
    finally:
        workbook.close()


def _batch_update_credit_rating_colors(excel_path: str, dry_run: bool = False) -> dict:
    workbook = load_workbook(filename=excel_path)
    update_count = 0
    today = datetime.now().date()
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(min_row=2):
                label_cell = row[0]
                if not (label_cell.value and "신용평가" in str(label_cell.value)):
                    continue

                for cell in row[1:]:
                    value_text = str(cell.value or "").strip()
                    if value_text == "":
                        cell.fill = copy(NO_FILL)
                        update_count += 1
                        continue

                    match = re.search(r"~(\d{2,4}\.\d{2}\.\d{2})", value_text)
                    if not match:
                        continue
                    expiry = _parse_expiry_date(match.group(1))
                    if expiry is None:
                        continue

                    cell.fill = copy(BLUE_FILL if expiry < today else GREEN_FILL)
                    update_count += 1

        if not dry_run:
            workbook.save(excel_path)
        return {"updatedCount": update_count, "dryRun": dry_run}
    finally:
        workbook.close()


@app.get("/health", response_model=ApiResponse)
def health() -> ApiResponse:
    return ApiResponse(
        success=True,
        message="ok",
        data={
            "service": APP_NAME,
            "time": datetime.now(timezone.utc).isoformat(),
            "storageRoot": str(DEFAULT_STORAGE_ROOT),
            "archiveRoot": str(_resolve_archive_root()),
            "datasetRoot": str(_resolve_dataset_root()),
            "dbPaths": _resolve_db_paths(),
        },
    )


@app.post("/excel-edit/upload", response_model=ApiResponse)
async def upload_excel_edit_files(
    files: list[UploadFile] = File(...),
    fileType: str = Form(default="전기경영상태"),
) -> ApiResponse:
    if not files:
        raise HTTPException(status_code=400, detail="업로드할 파일이 없습니다.")

    # 미리보기 업로드는 서버 디스크에 저장하지 않고 메모리로만 검증한다.
    previews = []
    for item in files:
        content = await item.read()
        safe_name = _safe_filename(item.filename or "unknown")
        previews.append(
            {
                "originalName": item.filename,
                "savedName": safe_name,
                "size": len(content),
                "contentType": item.content_type or "",
                "stored": False,
            }
        )

    return ApiResponse(
        success=True,
        message="미리보기 파일 업로드를 처리했습니다. (서버 저장 없음)",
        data={
            "fileType": fileType,
            "stored": False,
            "files": previews,
        },
    )


@app.post("/excel-edit/company-lookup", response_model=ApiResponse)
def company_lookup(payload: LookupRequest) -> ApiResponse:
    biz_no = payload.bizNo.strip()
    if not biz_no:
        raise HTTPException(status_code=400, detail="사업자등록번호가 필요합니다.")

    db_paths = _resolve_db_paths()
    if not db_paths:
        raise HTTPException(status_code=400, detail="DB 경로가 설정되지 않았습니다. 환경변수를 확인하세요.")

    if payload.fileType == "신용평가":
        candidates = [(db_type, path) for db_type, path in db_paths.items() if Path(path).exists()]
    else:
        target_db_type = FILE_TYPE_TO_DB_KEY.get(payload.fileType)
        if not target_db_type:
            raise HTTPException(status_code=400, detail=f"지원하지 않는 자료 종류입니다: {payload.fileType}")
        target_path = db_paths.get(target_db_type, "")
        candidates = [(target_db_type, target_path)] if target_path else []

    for db_type, excel_path in candidates:
        if not excel_path or not Path(excel_path).exists():
            continue
        position = _find_company_position(excel_path, biz_no)
        if not position:
            continue
        sheet_name, row, col = position
        raw = _extract_company_data(excel_path, sheet_name, row, col)
        raw_cells = _extract_company_cells(excel_path, sheet_name, row, col)
        lookup_payload = _build_lookup_payload(raw, db_type, excel_path, raw_cells=raw_cells)
        lookup_payload["version"] = _build_found_version(db_type, excel_path, biz_no, sheet_name, row, col, raw)
        return ApiResponse(
            success=True,
            message="업체 조회가 완료되었습니다.",
            data={"found": True, **lookup_payload},
        )

    return ApiResponse(
        success=True,
        message="업체를 찾지 못했습니다.",
        data={"found": False, "version": _build_not_found_version(payload.fileType, biz_no)},
    )


@app.post("/excel-edit/save", response_model=ApiResponse)
async def save_excel_edit_data(
    payload: str = Form(...),
    files: list[UploadFile] = File(default=[]),
) -> ApiResponse:
    try:
        request = SaveRequest.model_validate(json.loads(payload))
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"저장 요청 파라미터가 올바르지 않습니다: {exc}") from exc

    save_mode = str(request.saveMode or "normal").strip() or "normal"
    if save_mode not in {"normal", "archive_only"}:
        raise HTTPException(status_code=400, detail=f"지원하지 않는 저장 모드입니다: {save_mode}")

    biz_no = request.bizNo.strip()
    db_paths = _resolve_db_paths()
    expected_version = str(request.expectedVersion or "").strip()

    updated: dict = {}
    company_name = str(request.data.get("companyName") or "")
    region_name = str(request.data.get("region") or "")

    if save_mode == "archive_only":
        if not files:
            raise HTTPException(status_code=400, detail="파일만 저장 모드는 업로드 파일이 필요합니다.")

        if (not company_name or not region_name) and biz_no:
            for db_type, excel_path in db_paths.items():
                if not excel_path or not Path(excel_path).exists():
                    continue
                position = _find_company_position(excel_path, biz_no)
                if not position:
                    continue
                sheet_name, row, col = position
                raw = _extract_company_data(excel_path, sheet_name, row, col)
                company_name = company_name or str(raw.get("상호") or "")
                region_name = region_name or str(raw.get("지역") or "")
                if company_name and region_name:
                    break

        if not company_name:
            raise HTTPException(status_code=400, detail="파일만 저장에는 업체명이 필요합니다.")
        if not region_name:
            raise HTTPException(status_code=400, detail="파일만 저장에는 지역이 필요합니다.")

        if expected_version and db_paths:
            target_paths = [path for path in db_paths.values() if path and Path(path).exists()]
            with _exclusive_excel_locks(target_paths):
                _validate_expected_version_for_save(request, db_paths, expected_version)

        archived_files = _archive_uploaded_files(files, company_name, request.fileType, region_name)
        return ApiResponse(
            success=True,
            message="파일 보관이 완료되었습니다.",
            data={
                "saveMode": save_mode,
                "updated": updated,
                "archivedFiles": archived_files,
                "companyName": company_name,
                "region": region_name,
            },
        )

    if not biz_no:
        raise HTTPException(status_code=400, detail="사업자등록번호가 필요합니다.")
    if not db_paths:
        raise HTTPException(status_code=400, detail="DB 경로가 설정되지 않았습니다. 환경변수를 확인하세요.")

    if request.fileType == "신용평가":
        target_paths = [path for _, path in db_paths.items() if path and Path(path).exists()]
        with _exclusive_excel_locks(target_paths):
            _validate_expected_version_for_save(request, db_paths, expected_version)
            credit_text = _build_credit_text(request.data)
            results = _update_credit_data(db_paths, biz_no, credit_text)
            if not any(item.get("updated") for item in results):
                raise HTTPException(status_code=404, detail="신용평가 갱신 대상 업체를 찾지 못했습니다.")
            updated["credit"] = results
            if not company_name or not region_name:
                for db_type, path in db_paths.items():
                    position = _find_company_position(path, biz_no)
                    if not position:
                        continue
                    sheet_name, row, col = position
                    raw = _extract_company_data(path, sheet_name, row, col)
                    company_name = company_name or str(raw.get("상호") or "")
                    region_name = region_name or str(raw.get("지역") or "")
                    break
    else:
        db_key = FILE_TYPE_TO_DB_KEY.get(request.fileType)
        if not db_key:
            raise HTTPException(status_code=400, detail=f"지원하지 않는 자료 종류입니다: {request.fileType}")
        excel_path = db_paths.get(db_key, "")
        if not excel_path or not Path(excel_path).exists():
            raise HTTPException(status_code=400, detail=f"{db_key} DB 경로가 유효하지 않습니다.")
        with _exclusive_excel_lock(excel_path):
            _validate_expected_version_for_save(request, db_paths, expected_version)
            position = _find_company_position(excel_path, biz_no)
            if position:
                updated_labels = _update_management_data(excel_path, biz_no, request.data, db_key)
                updated["management"] = {
                    "dbType": db_key,
                    "excelPath": excel_path,
                    "action": "update_existing",
                    "updatedLabels": updated_labels,
                }
                if not company_name or not region_name:
                    sheet_name, row, col = position
                    raw = _extract_company_data(excel_path, sheet_name, row, col)
                    company_name = company_name or str(raw.get("상호") or "")
                    region_name = region_name or str(raw.get("지역") or "")
            else:
                added = _add_new_company_data(excel_path, request.data, db_key)
                company_name = company_name or str(added.get("companyName") or "")
                region_name = region_name or str(added.get("sheetName") or "")
                updated["management"] = {
                    "dbType": db_key,
                    "excelPath": excel_path,
                    "action": "add_new_company",
                    "sheetName": region_name,
                    "companyName": company_name,
                }

    archived_files = _archive_uploaded_files(files, company_name, request.fileType, region_name) if files else []

    return ApiResponse(
        success=True,
        message="확정 및 저장이 완료되었습니다.",
        data={
            "saveMode": save_mode,
            "updated": updated,
            "archivedFiles": archived_files,
            "companyName": company_name,
            "region": region_name,
        },
    )


@app.post("/excel-edit/delete-company", response_model=ApiResponse)
def delete_excel_edit_company(payload: DeleteRequest) -> ApiResponse:
    biz_no = payload.bizNo.strip()
    if not biz_no:
        raise HTTPException(status_code=400, detail="사업자등록번호가 필요합니다.")

    if payload.fileType == "신용평가":
        raise HTTPException(status_code=400, detail="신용평가 삭제는 아직 지원하지 않습니다.")

    db_paths = _resolve_db_paths()
    if not db_paths:
        raise HTTPException(status_code=400, detail="DB 경로가 설정되지 않았습니다. 환경변수를 확인하세요.")

    db_key = FILE_TYPE_TO_DB_KEY.get(payload.fileType)
    if not db_key:
        raise HTTPException(status_code=400, detail=f"지원하지 않는 자료 종류입니다: {payload.fileType}")

    excel_path = db_paths.get(db_key, "")
    if not excel_path or not Path(excel_path).exists():
        raise HTTPException(status_code=400, detail=f"{db_key} DB 경로가 유효하지 않습니다.")

    with _exclusive_excel_lock(excel_path):
        _validate_expected_version_for_save(
            SaveRequest(fileType=payload.fileType, bizNo=biz_no, expectedVersion=payload.expectedVersion),
            db_paths,
            str(payload.expectedVersion or "").strip(),
        )
        deleted = _delete_management_company(excel_path, biz_no, db_key)

    return ApiResponse(
        success=True,
        message="업체 삭제가 완료되었습니다.",
        data={
            "deleted": deleted,
            "fileType": payload.fileType,
            "bizNo": _format_biz_no(biz_no),
        },
    )


@app.post("/excel-edit/update-year-end-color", response_model=ApiResponse)
def update_year_end_color(payload: JobRequest) -> ApiResponse:
    targets = _resolve_target_excel_paths_for_job(payload)
    results: list[dict] = []
    total = 0
    with _exclusive_excel_locks([path for _, path in targets]):
        for db_type, excel_path in targets:
            result = _batch_update_year_end_colors(excel_path, dry_run=payload.dryRun)
            total += int(result.get("updatedCount", 0))
            results.append({"dbType": db_type, "excelPath": excel_path, **result})

    return ApiResponse(
        success=True,
        message=f"연말 색상 업데이트 완료 ({total}개 셀)",
        data={
            "implemented": True,
            "request": payload.model_dump(),
            "results": results,
            "updatedCount": total,
        },
    )


@app.post("/excel-edit/update-credit-expiry", response_model=ApiResponse)
def update_credit_expiry(payload: JobRequest) -> ApiResponse:
    targets = _resolve_target_excel_paths_for_job(payload)
    results: list[dict] = []
    total = 0
    with _exclusive_excel_locks([path for _, path in targets]):
        for db_type, excel_path in targets:
            result = _batch_update_credit_rating_colors(excel_path, dry_run=payload.dryRun)
            total += int(result.get("updatedCount", 0))
            results.append({"dbType": db_type, "excelPath": excel_path, **result})

    return ApiResponse(
        success=True,
        message=f"신용평가 유효기간 갱신 완료 ({total}개 셀)",
        data={
            "implemented": True,
            "request": payload.model_dump(),
            "results": results,
            "updatedCount": total,
        },
    )


@app.post("/excel-edit/render-pdf-page")
async def render_pdf_page(
    file: UploadFile = File(...),
    page: int = Form(default=1),
    dpi: int = Form(default=160),
) -> Response:
    if not file:
        raise HTTPException(status_code=400, detail="PDF 파일이 필요합니다.")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="빈 파일입니다.")

    if dpi < 72:
        dpi = 72
    if dpi > 300:
        dpi = 300

    try:
        with fitz.open(stream=content, filetype="pdf") as doc:
            page_count = doc.page_count
            if page_count <= 0:
                raise HTTPException(status_code=400, detail="페이지가 없는 PDF입니다.")

            target_page = max(1, min(page, page_count))
            pdf_page = doc.load_page(target_page - 1)
            pix = pdf_page.get_pixmap(dpi=dpi, alpha=False)
            image_bytes = pix.tobytes("png")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"PDF 렌더링 실패: {exc}") from exc

    return Response(
        content=image_bytes,
        media_type="image/png",
        headers={
            "x-pdf-page-count": str(page_count),
            "x-pdf-page-number": str(target_page),
        },
    )


@app.post("/excel-edit/render-image")
async def render_image(file: UploadFile = File(...)) -> Response:
    if not file:
        raise HTTPException(status_code=400, detail="이미지 파일이 필요합니다.")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="빈 파일입니다.")

    try:
        with Image.open(BytesIO(content)) as img:
            converted = img.convert("RGB")
            output = BytesIO()
            converted.save(output, format="PNG", optimize=True)
            image_bytes = output.getvalue()
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"이미지 렌더링 실패: {exc}") from exc

    return Response(
        content=image_bytes,
        media_type="image/png",
    )


@app.post("/excel-edit/export-pdf-pages")
async def export_pdf_pages(
    file: UploadFile = File(...),
    pages: str = Form(default=""),
) -> Response:
    if not file:
        raise HTTPException(status_code=400, detail="PDF 파일이 필요합니다.")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="빈 파일입니다.")

    try:
        with fitz.open(stream=content, filetype="pdf") as source:
            page_count = source.page_count
            if page_count <= 0:
                raise HTTPException(status_code=400, detail="페이지가 없는 PDF입니다.")

            selected_pages = _parse_pdf_page_selection(pages, page_count)
            target = fitz.open()
            try:
                for page_index in selected_pages:
                    target.insert_pdf(source, from_page=page_index, to_page=page_index)
                exported_bytes = target.tobytes(garbage=3, deflate=True)
            finally:
                target.close()
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"PDF 내보내기 실패: {exc}") from exc

    return Response(
        content=exported_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": 'attachment; filename="exported-pages.pdf"',
            "x-pdf-page-count": str(len(selected_pages)),
        },
    )


@app.post("/excel-edit/remove-pdf-pages")
async def remove_pdf_pages(
    file: UploadFile = File(...),
    pages: str = Form(default=""),
) -> Response:
    if not file:
        raise HTTPException(status_code=400, detail="PDF 파일이 필요합니다.")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="빈 파일입니다.")

    try:
        with fitz.open(stream=content, filetype="pdf") as source:
            page_count = source.page_count
            if page_count <= 0:
                raise HTTPException(status_code=400, detail="페이지가 없는 PDF입니다.")

            selected_pages = set(_parse_pdf_page_selection(pages, page_count))
            remain_pages = [idx for idx in range(page_count) if idx not in selected_pages]
            if not remain_pages:
                return Response(status_code=204, headers={"x-pdf-page-count": "0"})

            target = fitz.open()
            try:
                for page_index in remain_pages:
                    target.insert_pdf(source, from_page=page_index, to_page=page_index)
                remain_bytes = target.tobytes(garbage=3, deflate=True)
            finally:
                target.close()
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"PDF 페이지 삭제 실패: {exc}") from exc

    return Response(
        content=remain_bytes,
        media_type="application/pdf",
        headers={
            "x-pdf-page-count": str(len(remain_pages)),
        },
    )
