import math
from pathlib import Path

from openpyxl import load_workbook

from text_utils import extract_manager_name, normalize_name

_DB_CACHE = {"path": None, "mtime": None, "data": []}


def _to_number(val):
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return None
    s = str(val).replace(",", "").strip()
    if s == "":
        return None
    try:
        return float(s)
    except Exception:
        return None


def _load_sheet_entries(ws, sheet_name, relative_offsets):
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    merged_value = {}
    for merged in ws.merged_cells.ranges:
        tl = ws.cell(merged.min_row, merged.min_col).value
        for r in range(merged.min_row, merged.max_row + 1):
            for c in range(merged.min_col, merged.max_col + 1):
                merged_value[(r, c)] = tl

    def get_value(r, c):
        v = ws.cell(r, c).value
        if v is None:
            return merged_value.get((r, c))
        return v

    header_positions = []
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = get_value(r, c)
            if cell is None:
                continue
            val = str(cell)
            if "회사명" in val:
                header_positions.append((r, c))

    if not header_positions:
        return []

    entries = []
    seen_keys = set()
    for header_row, header_col in header_positions:
        for col in range(header_col + 1, max_col + 1):
            raw_name = get_value(header_row, col)
            if raw_name is None:
                continue
            raw_name = str(raw_name).strip()
            if not raw_name:
                continue
            name = raw_name.split("\n")[0].strip()
            if not name:
                continue
            dedup_key = (sheet_name, header_row, col, name)
            if dedup_key in seen_keys:
                continue
            seen_keys.add(dedup_key)
            entry = {
                "name": name,
                "norm": normalize_name(name),
                "region": sheet_name.strip(),
                "bizNo": "",
                "debtRatio": None,
                "currentRatio": None,
                "bizYears": None,
                "perf5y": None,
                "creditGrade": "",
                "sipyung": None,
                "notes": "",
                "managerName": "",
            }
            for key, offset in relative_offsets.items():
                r = header_row + offset
                if r > max_row:
                    continue
                val = get_value(r, col)
                if key in {"부채비율", "유동비율"} and isinstance(val, (int, float)):
                    val = val * 100
                if key == "사업자번호":
                    entry["bizNo"] = "" if val is None else str(val).strip()
                elif key == "부채비율":
                    entry["debtRatio"] = _to_number(val)
                elif key == "유동비율":
                    entry["currentRatio"] = _to_number(val)
                elif key == "영업기간":
                    entry["bizYears"] = _to_number(val)
                elif key == "시평":
                    entry["sipyung"] = _to_number(val)
                elif key == "5년 실적":
                    entry["perf5y"] = _to_number(val)
                elif key == "신용평가":
                    entry["creditGrade"] = "" if val is None else str(val).strip()
                elif key == "비고":
                    entry["notes"] = "" if val is None else str(val).strip()
            entry["managerName"] = extract_manager_name(entry.get("notes", ""))
            entries.append(entry)
    return entries


def load_db(db_path: Path):
    wb = load_workbook(db_path, data_only=False)
    data = []
    relative_offsets = {
        "대표자": 1,
        "사업자번호": 2,
        "지역": 3,
        "시평": 4,
        "3년 실적": 5,
        "5년 실적": 6,
        "부채비율": 7,
        "유동비율": 8,
        "영업기간": 9,
        "신용평가": 10,
        "여성기업": 11,
        "중소기업": 12,
        "일자리창출": 13,
        "품질평가": 14,
        "비고": 15,
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        entries = _load_sheet_entries(ws, sheet_name, relative_offsets)
        data.extend(entries)
    return data


def load_db_cached(db_path: Path, force=False):
    mtime = db_path.stat().st_mtime if db_path.exists() else None
    if (
        not force
        and _DB_CACHE["path"] == str(db_path)
        and _DB_CACHE["mtime"] == mtime
    ):
        return _DB_CACHE["data"]
    data = load_db(db_path)
    _DB_CACHE["path"] = str(db_path)
    _DB_CACHE["mtime"] = mtime
    _DB_CACHE["data"] = data
    return data


def load_db_stats(db_path: Path):
    wb = load_workbook(db_path, data_only=False)
    relative_offsets = {
        "대표자": 1,
        "사업자번호": 2,
        "지역": 3,
        "시평": 4,
        "3년 실적": 5,
        "5년 실적": 6,
        "부채비율": 7,
        "유동비율": 8,
        "영업기간": 9,
        "신용평가": 10,
        "여성기업": 11,
        "중소기업": 12,
        "일자리창출": 13,
        "품질평가": 14,
        "비고": 15,
    }
    stats = []
    total = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        entries = _load_sheet_entries(ws, sheet_name, relative_offsets)
        count = len(entries)
        total += count
        if count:
            stats.append((sheet_name, count))
    return total, stats
